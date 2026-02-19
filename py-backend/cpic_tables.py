"""
CPIC Table Loader — Gene-Agnostic Auto-Discovery
==================================================
Scans ``data/tables/`` for official CPIC Excel files and loads them
automatically for **every gene** found.  Expected file-naming convention::

    {GENE}_allele_definition_table.xlsx
    {GENE}_allele_functionality_reference.xlsx
    {GENE}_Diplotype_Phenotype_Table.xlsx

Not every gene needs all three files.  If only two are present the
loader will still pick up what it can.

All data is loaded once at import time and stored in per-gene registries
(plain dicts) so the rest of the codebase can call gene-agnostic helpers
like ``get_activity_value("CYP2D6", "*4")``.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

# ---------------------------------------------------------------------------
# Base path
# ---------------------------------------------------------------------------
_DATA_DIR = Path(__file__).resolve().parent / "data" / "tables"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _allele_sort_key(allele: str) -> float:
    """Best-effort numeric sort for star-allele names (*1 < *2 < *10 …)."""
    a = allele.lstrip("*").split("x")[0].split("+")[0]
    a = a.replace("A", ".1").replace("B", ".2").replace("C", ".3")
    try:
        return float(a)
    except ValueError:
        return 999.0


# ═══════════════════════════════════════════════════════════════════════════
# Data-class wrappers (unchanged API from before)
# ═══════════════════════════════════════════════════════════════════════════

class AlleleFunction:
    """One row of an Allele Functionality Reference table."""
    __slots__ = ("allele", "activity_value", "clinical_function", "evidence")

    def __init__(self, allele: str, activity_value: Optional[float],
                 clinical_function: str, evidence: str):
        self.allele = allele
        self.activity_value = activity_value
        self.clinical_function = clinical_function
        self.evidence = evidence

    def to_legacy_function(self) -> str:
        cf = self.clinical_function.lower()
        if "no function" in cf:
            return "no_function"
        elif "decreased" in cf:
            return "decreased"
        elif "increased" in cf:
            return "increased"
        elif "normal" in cf:
            return "normal"
        elif "uncertain" in cf or "unknown" in cf:
            return "unknown"
        return "normal"

    def __repr__(self) -> str:
        return f"AlleleFunction({self.allele}, AV={self.activity_value}, {self.clinical_function})"


class DiplotypePhenotype:
    """One row of a Diplotype-Phenotype table."""
    __slots__ = ("diplotype", "activity_score", "phenotype", "ehr_priority",
                 "_gene")

    def __init__(self, diplotype: str, activity_score: Optional[float],
                 phenotype: str, ehr_priority: str, gene: str = ""):
        self.diplotype = diplotype
        self.activity_score = activity_score
        self.phenotype = phenotype
        self.ehr_priority = ehr_priority
        self._gene = gene

    @property
    def metabolizer_phenotype(self) -> str:
        """Strip the gene prefix (e.g. 'CYP2D6 Normal Metabolizer' → 'Normal Metabolizer')."""
        p = self.phenotype
        if self._gene and p.upper().startswith(self._gene.upper() + " "):
            p = p[len(self._gene) + 1:]
        return p

    @property
    def is_high_risk(self) -> bool:
        return "High Risk" in (self.ehr_priority or "")

    def __repr__(self) -> str:
        return f"DiplotypePhenotype({self.diplotype}, AS={self.activity_score}, {self.phenotype})"


# ═══════════════════════════════════════════════════════════════════════════
# Per-gene registries — populated by _discover_and_load()
# ═══════════════════════════════════════════════════════════════════════════

# { gene: { rsid: (gene, star_allele) } }
GENE_RSID_TO_ALLELE: Dict[str, Dict[str, Tuple[str, str]]] = {}

# { gene: { star_allele: [rsids] } }
GENE_ALLELE_TO_RSIDS: Dict[str, Dict[str, List[str]]] = {}

# { gene: { star_allele: AlleleFunction } }
GENE_ALLELE_FUNCTION: Dict[str, Dict[str, AlleleFunction]] = {}

# { gene: { diplotype_str: DiplotypePhenotype } }  (includes reverse keys)
GENE_DIPLOTYPE_PHENOTYPE: Dict[str, Dict[str, DiplotypePhenotype]] = {}

# Set of genes that were successfully loaded from CPIC tables
LOADED_GENES: Set[str] = set()


# ═══════════════════════════════════════════════════════════════════════════
# 1.  Allele Definition Table loader (gene-agnostic)
# ═══════════════════════════════════════════════════════════════════════════

def _load_allele_definitions(filepath: Path, gene: str) -> Tuple[
    Dict[str, Tuple[str, str]],
    Dict[str, List[str]],
]:
    rsid_to_allele: Dict[str, Tuple[str, str]] = {}
    allele_to_rsids: Dict[str, List[str]] = {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    # Try common sheet names
    ws = None
    for name in ("Alleles", "Sheet1", wb.sheetnames[0]):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return rsid_to_allele, allele_to_rsids

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 7:
        return rsid_to_allele, allele_to_rsids

    # Row 5 (0-indexed) = rsIDs across columns
    rsid_row = rows[5]
    col_rsids: Dict[int, str] = {}
    for ci in range(1, len(rsid_row)):
        val = rsid_row[ci]
        if val and str(val).strip().startswith("rs"):
            for rs in str(val).strip().split(";"):
                rs = rs.strip()
                if rs.startswith("rs"):
                    col_rsids[ci] = rs

    # First pass: find *1 (reference) columns to know what is reference data
    for ri in range(6, len(rows)):
        row = rows[ri]
        allele_name = row[0]
        if not allele_name:
            continue
        if str(allele_name).strip() == "*1":
            break

    # Collect non-reference alleles
    _allele_defining: Dict[str, List[str]] = {}

    # Build a label to skip (e.g. "CYP2D6 Allele", "CYP2C19 Allele")
    allele_header_label = f"{gene} Allele"

    for ri in range(6, len(rows)):
        row = rows[ri]
        allele_name = row[0]
        if not allele_name:
            continue
        allele_name = str(allele_name).strip()

        if allele_name in ("*1", allele_header_label):
            continue
        if not allele_name.startswith("*"):
            continue

        allele_rsids_list: List[str] = []
        for ci, rsid in col_rsids.items():
            if ci < len(row) and row[ci] is not None:
                cell_val = str(row[ci]).strip()
                if cell_val and cell_val not in ("", "None"):
                    allele_rsids_list.append(rsid)

        if allele_rsids_list:
            _allele_defining[allele_name] = allele_rsids_list
            allele_to_rsids[allele_name] = allele_rsids_list

    # Build rsid→allele preferring lowest-numbered allele
    sorted_alleles = sorted(_allele_defining.keys(), key=_allele_sort_key)
    for allele in sorted_alleles:
        for rsid in _allele_defining[allele]:
            if rsid not in rsid_to_allele:
                rsid_to_allele[rsid] = (gene, allele)

    return rsid_to_allele, allele_to_rsids


# ═══════════════════════════════════════════════════════════════════════════
# 2.  Allele Functionality Reference loader (gene-agnostic)
# ═══════════════════════════════════════════════════════════════════════════

def _load_allele_functionality(filepath: Path) -> Dict[str, AlleleFunction]:
    result: Dict[str, AlleleFunction] = {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for name in ("Allele Function", "Sheet1", wb.sheetnames[0]):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return result

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 0 = gene header, Row 1 = column headers, Row 2+ = data
    for ri in range(2, len(rows)):
        row = rows[ri]
        allele_raw = row[0]
        if not allele_raw:
            continue
        allele = str(allele_raw).strip()

        av_raw = row[1]
        activity_value: Optional[float] = None
        if av_raw is not None:
            av_str = str(av_raw).strip().replace("≥", "").replace(">=", "")
            try:
                activity_value = float(av_str)
            except ValueError:
                activity_value = None

        clinical_function = str(row[3]).strip() if row[3] else "Unknown function"
        evidence = str(row[6]).strip() if len(row) > 6 and row[6] else "Unknown"

        result[allele] = AlleleFunction(allele, activity_value, clinical_function, evidence)

    return result


# ═══════════════════════════════════════════════════════════════════════════
# 3.  Diplotype-Phenotype Table loader (gene-agnostic)
# ═══════════════════════════════════════════════════════════════════════════

def _load_diplotype_phenotype(filepath: Path, gene: str) -> Dict[str, DiplotypePhenotype]:
    result: Dict[str, DiplotypePhenotype] = {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for name in ("Diplotypes", "Sheet1", wb.sheetnames[0]):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return result

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    for ri in range(1, len(rows)):
        row = rows[ri]
        diplotype_raw = row[0]
        if not diplotype_raw:
            continue
        diplotype = str(diplotype_raw).strip()

        as_raw = row[1]
        activity_score: Optional[float] = None
        if as_raw is not None:
            as_str = str(as_raw).strip().replace("≥", "").replace(">=", "")
            try:
                activity_score = float(as_str)
            except ValueError:
                activity_score = None

        phenotype = str(row[2]).strip() if row[2] else f"{gene} Indeterminate"
        ehr_priority = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        result[diplotype] = DiplotypePhenotype(
            diplotype, activity_score, phenotype, ehr_priority, gene
        )

        # Store reverse order too
        parts = diplotype.split("/")
        if len(parts) == 2:
            reverse_key = f"{parts[1]}/{parts[0]}"
            if reverse_key not in result:
                result[reverse_key] = DiplotypePhenotype(
                    reverse_key, activity_score, phenotype, ehr_priority, gene
                )

    return result


# ═══════════════════════════════════════════════════════════════════════════
# Auto-discovery & loading
# ═══════════════════════════════════════════════════════════════════════════

# File-name patterns  (case-insensitive matching)
_PAT_ALLELE_DEF  = re.compile(r"^(.+?)_allele_definition_table\.xlsx$", re.IGNORECASE)
_PAT_ALLELE_FUNC = re.compile(r"^(.+?)_allele_functionality_reference\.xlsx$", re.IGNORECASE)
_PAT_DIPLO_PHENO = re.compile(r"^(.+?)_Diplotype_Phenotype_Table\.xlsx$", re.IGNORECASE)


def _discover_and_load() -> None:
    """Scan _DATA_DIR for CPIC Excel files and load them into the registries."""
    if not _DATA_DIR.exists():
        print(f"[cpic_tables] WARNING: {_DATA_DIR} does not exist — no CPIC tables loaded")
        return

    # Group files by gene
    gene_files: Dict[str, Dict[str, Path]] = {}  # gene → {"def"|"func"|"diplo": path}

    for f in sorted(_DATA_DIR.iterdir()):
        if not f.is_file() or not f.suffix.lower() == ".xlsx":
            continue

        m = _PAT_ALLELE_DEF.match(f.name)
        if m:
            gene = m.group(1).upper()
            gene_files.setdefault(gene, {})["def"] = f
            continue

        m = _PAT_ALLELE_FUNC.match(f.name)
        if m:
            gene = m.group(1).upper()
            gene_files.setdefault(gene, {})["func"] = f
            continue

        m = _PAT_DIPLO_PHENO.match(f.name)
        if m:
            gene = m.group(1).upper()
            gene_files.setdefault(gene, {})["diplo"] = f

    if not gene_files:
        print("[cpic_tables] No CPIC Excel tables found in", _DATA_DIR)
        return

    print(f"[cpic_tables] Discovered CPIC tables for {len(gene_files)} gene(s): "
          f"{', '.join(sorted(gene_files.keys()))}")

    for gene in sorted(gene_files.keys()):
        files = gene_files[gene]
        print(f"\n  ── {gene} ──")

        # 1. Allele definitions
        if "def" in files:
            rsid_map, allele_map = _load_allele_definitions(files["def"], gene)
            GENE_RSID_TO_ALLELE[gene] = rsid_map
            GENE_ALLELE_TO_RSIDS[gene] = allele_map
            print(f"    Allele Definition:   {len(rsid_map)} rsID mappings, "
                  f"{len(allele_map)} alleles")

        # 2. Allele functionality
        if "func" in files:
            func_map = _load_allele_functionality(files["func"])
            GENE_ALLELE_FUNCTION[gene] = func_map
            print(f"    Allele Functionality: {len(func_map)} entries")

        # 3. Diplotype-phenotype
        if "diplo" in files:
            diplo_map = _load_diplotype_phenotype(files["diplo"], gene)
            raw_count = sum(1 for k, v in diplo_map.items() if k == v.diplotype)
            GENE_DIPLOTYPE_PHENOTYPE[gene] = diplo_map
            print(f"    Diplotype-Phenotype:  {raw_count} diplotypes "
                  f"({len(diplo_map)} incl. reverse)")

        LOADED_GENES.add(gene)

    print(f"\n[cpic_tables] Done — {len(LOADED_GENES)} gene(s) loaded.\n")


# Run auto-discovery at import time
_discover_and_load()

# Legacy aliases for backward compat (point to CYP2D6 if available)
CYP2D6_RSID_TO_ALLELE  = GENE_RSID_TO_ALLELE.get("CYP2D6", {})
CYP2D6_ALLELE_TO_RSIDS = GENE_ALLELE_TO_RSIDS.get("CYP2D6", {})
CYP2D6_ALLELE_FUNCTION = GENE_ALLELE_FUNCTION.get("CYP2D6", {})
CYP2D6_DIPLOTYPE_PHENOTYPE = GENE_DIPLOTYPE_PHENOTYPE.get("CYP2D6", {})


# ═══════════════════════════════════════════════════════════════════════════
# Public API — gene-agnostic convenience functions
# ═══════════════════════════════════════════════════════════════════════════

def has_gene(gene: str) -> bool:
    """Return True if CPIC tables were loaded for *gene*."""
    return gene.upper() in LOADED_GENES


def loaded_genes() -> List[str]:
    """Return sorted list of genes with loaded CPIC tables."""
    return sorted(LOADED_GENES)


def lookup_rsid(rsid: str, gene: Optional[str] = None) -> Optional[Tuple[str, str]]:
    """
    Look up an rsID → (gene, star_allele).

    If *gene* is provided, search only that gene's table.
    Otherwise search all loaded genes (first match wins,
    deterministic order).
    """
    if gene:
        g = gene.upper()
        return GENE_RSID_TO_ALLELE.get(g, {}).get(rsid)
    # Search all genes in deterministic order
    for g in sorted(GENE_RSID_TO_ALLELE.keys()):
        hit = GENE_RSID_TO_ALLELE[g].get(rsid)
        if hit:
            return hit
    return None


def get_allele_function(gene: str, allele: str) -> Optional[AlleleFunction]:
    """Get the AlleleFunction entry for a gene + star allele."""
    return GENE_ALLELE_FUNCTION.get(gene.upper(), {}).get(allele)


def get_activity_value(gene: str, allele: str) -> float:
    """CPIC activity value for an allele.  Defaults to 1.0 if unknown."""
    af = GENE_ALLELE_FUNCTION.get(gene.upper(), {}).get(allele)
    if af and af.activity_value is not None:
        return af.activity_value
    return 1.0


def get_clinical_function(gene: str, allele: str) -> str:
    """Clinical function label (e.g. 'No function') for a star allele."""
    af = GENE_ALLELE_FUNCTION.get(gene.upper(), {}).get(allele)
    if af:
        return af.clinical_function
    return "Unknown function"


def lookup_diplotype_phenotype(gene: str, diplotype: str) -> Optional[DiplotypePhenotype]:
    """Look up a diplotype in the official CPIC table for *gene*."""
    return GENE_DIPLOTYPE_PHENOTYPE.get(gene.upper(), {}).get(diplotype)


def infer_phenotype_from_diplotype(gene: str, diplotype: str) -> Optional[str]:
    """
    Returns the metabolizer phenotype string (e.g. 'Intermediate Metabolizer')
    from the CPIC table, or None if not found.
    """
    dp = GENE_DIPLOTYPE_PHENOTYPE.get(gene.upper(), {}).get(diplotype)
    if dp:
        return dp.metabolizer_phenotype
    return None


def get_rsids_for_allele(gene: str, allele: str) -> List[str]:
    """Return all rsIDs that define *allele* for *gene*."""
    return GENE_ALLELE_TO_RSIDS.get(gene.upper(), {}).get(allele, [])


# ═══════════════════════════════════════════════════════════════════════════
# Legacy-compatible builders  (used by pgx_knowledgebase.py)
# ═══════════════════════════════════════════════════════════════════════════

def build_legacy_allele_function_dict(gene: str) -> Dict[str, str]:
    """
    Build ``{ "*4": "no_function", … }`` for a gene, compatible with the
    ``ALLELE_FUNCTION[gene]`` format in pgx_knowledgebase.
    """
    result: Dict[str, str] = {}
    for allele, af in GENE_ALLELE_FUNCTION.get(gene.upper(), {}).items():
        if "x" in allele:
            continue
        result[allele] = af.to_legacy_function()
    return result


def build_legacy_rsid_to_allele_dict(gene: str) -> Dict[str, Tuple[str, str]]:
    """
    Build ``{ "rs3892097": ("CYP2D6", "*4"), … }`` for a gene,
    preferring the lowest-numbered allele when rsIDs are shared.
    """
    result: Dict[str, Tuple[str, str]] = {}
    g = gene.upper()
    allele_rsid_pairs: List[Tuple[str, str]] = []

    for allele, rsids in GENE_ALLELE_TO_RSIDS.get(g, {}).items():
        for rsid in rsids:
            allele_rsid_pairs.append((allele, rsid))

    allele_rsid_pairs.sort(key=lambda x: _allele_sort_key(x[0]))

    for allele, rsid in allele_rsid_pairs:
        if rsid not in result:
            result[rsid] = (g, allele)

    return result
